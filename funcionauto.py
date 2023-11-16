import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string

def automatizar_excel(nombre_archivo):
      """Input sales_mes.xlsx / Output report_mes.xlsx"""
        archivo_excel = pd.read_excel(nombre_archivo)
        #print(archivo_excel[['Gender', 'Product line', 'Total']])

        tabla_pivote = archivo_excel.pivot_table(index='Gender', columns= 'Product line', values='Total', aggfunc='sum').round(0)
        #print(tabla_pivote)

        mes_extension = nombre_archivo.split('_')[1]

        tabla_pivote.to_excel(f'sales_{mes_extension}', startrow=4, sheet_name='Report')

        wb = load_workbook(f'sales_{mes_extension}')
        pestana = wb['Report']

        min_col = wb.active.min_column
        max_col = wb.active.max_column
        min_fila = wb.active.min_row
        max_fila = wb.active.max_row

        #print(min_col)
        #print(max_col)
        #print(min_fila)
        #print(max_fila)

        #Grafico
        barchart = BarChart()
        data = Reference(pestana, min_col=min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila)
        categorias = Reference(pestana, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)
        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categorias)

        pestana.add_chart(barchart, 'B12')
        barchart.title = 'Ventas'
        barchart.style = 2

        #pestana['B8'] = '=sum(B6:B7)'
        #pestana['B8'].style = 'Currency'

        # pestana['C8'] = '=sum(B6:B7)'
        # pestana['C8'].style = 'Currency'

        abecedario = list(string.ascii_uppercase)
        #print(abecedario)
        abecedario_excel = abecedario[0:max_col]
        #print(abecedario_excel)

        for i in abecedario_excel:
            if i !='A':
                pestana[f'{i}{max_fila+1}'] = f'=sum({i}{min_fila+1}:{i}{max_fila})'
                pestana[f'{i}{max_fila+1}'].style = 'Currency'

        pestana[f'{abecedario_excel[0]}{max_fila+1}'] = 'Total'

        pestana['A1'] = 'Reporte'
        mes = mes_extension.split('.')[0]
        pestana['A2'] = mes

        pestana['A1'].font = Font('Arial', bold=True, size=20)
        pestana['A2'].font = Font('Arial', bold=True, size=12)

        wb.save(f'sales_{mes_extension}')
        return

#automatizar reporte 2021
#automatizar_excel('sales_2021.xlsx')

#automatizar reportes mensuales
automatizar_excel('sales_enero.xlsx')
#automatizar_excel('sales_febrero.xlsx')
#automatizar_excel('sales_marzo.xlsx')

